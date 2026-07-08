from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .artifacts import download_result, generated_dir, new_stored_name, safe_download_name
from .base import ToolContext, ToolDef, ToolResult


def _coerce(value):
    if isinstance(value, (int, float)):
        return value
    s = str(value)
    try:
        if s.strip().lstrip("-").isdigit():
            return int(s)
        return float(s)
    except (TypeError, ValueError):
        return s


class XlsxGenerateTool:
    definition = ToolDef(
        name="generate_xlsx",
        description=(
            "Create a downloadable Microsoft Excel (.xlsx) spreadsheet from structured "
            "tabular data that you author. Use this whenever the user asks for a "
            "spreadsheet, Excel file, or data table. Provide one or more sheets, each "
            "with column headers and rows of data. Numeric-looking values become real "
            "numbers. Returns a Markdown download link — include it verbatim in your "
            "reply."
        ),
        parameters={
            "type": "object",
            "properties": {
                "title": {"type": "string", "description": "Optional workbook title."},
                "sheets": {
                    "type": "array",
                    "description": "The worksheets, in order.",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {
                                "type": "string",
                                "description": "Sheet/tab name.",
                            },
                            "columns": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": "Header row labels.",
                            },
                            "rows": {
                                "type": "array",
                                "description": "Data rows (each an array of cell values).",
                                "items": {
                                    "type": "array",
                                    "items": {"type": ["string", "number"]},
                                },
                            },
                            "chart": {
                                "type": "object",
                                "description": (
                                    "Optional native chart built from this sheet's data."
                                ),
                                "properties": {
                                    "type": {
                                        "type": "string",
                                        "enum": ["bar", "column", "pie", "line"],
                                    },
                                    "title": {"type": "string"},
                                    "category_column": {
                                        "type": "integer",
                                        "description": "1-based column index for labels.",
                                    },
                                    "value_columns": {
                                        "type": "array",
                                        "items": {"type": "integer"},
                                        "description": "1-based column indices for values.",
                                    },
                                },
                            },
                        },
                        "required": ["columns", "rows"],
                    },
                },
            },
            "required": ["sheets"],
        },
    )

    def _add_chart(self, ws, spec: dict, ncols: int, nrows: int) -> None:
        try:
            ctype = (spec.get("type") or "column").strip().lower()
            cat_col = int(spec.get("category_column") or 1)
            value_cols = spec.get("value_columns") or list(range(2, ncols + 1))
            value_cols = [int(v) for v in value_cols if 1 <= int(v) <= ncols]
            if not value_cols:
                return
            if ctype == "pie":
                chart = PieChart()
                value_cols = value_cols[:1]
            elif ctype == "line":
                chart = LineChart()
            else:
                chart = BarChart()
                chart.type = "bar" if ctype == "bar" else "col"
            title = (spec.get("title") or "").strip()
            if title:
                chart.title = title
            cats = Reference(ws, min_col=cat_col, min_row=2, max_row=nrows + 1)
            for vc in value_cols:
                ref = Reference(ws, min_col=vc, min_row=1, max_row=nrows + 1)
                chart.add_data(ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 15
            anchor = f"{get_column_letter(ncols + 2)}2"
            ws.add_chart(chart, anchor)
        except Exception:  # noqa: BLE001
            return

    async def run(self, args: dict, ctx: ToolContext) -> ToolResult:
        title = (args.get("title") or "").strip() or "Workbook"
        sheets = args.get("sheets") or []
        if not isinstance(sheets, list) or not sheets:
            return ToolResult(content="No sheet data was provided.", citations=None)

        try:
            wb = Workbook()
            wb.remove(wb.active)
            header_fill = PatternFill("solid", fgColor="4F46E5")
            header_font = Font(bold=True, color="FFFFFF")

            count = 0
            for idx, sd in enumerate(sheets):
                if not isinstance(sd, dict):
                    continue
                count += 1
                name = (str(sd.get("name") or f"Sheet{idx + 1}"))[:31] or f"Sheet{idx + 1}"
                ws = wb.create_sheet(title=name)
                columns = [str(c) for c in (sd.get("columns") or [])]
                rows = sd.get("rows") or []

                widths: list[int] = [len(c) for c in columns]
                if columns:
                    ws.append(columns)
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(vertical="center")
                    ws.freeze_panes = "A2"

                for r in rows:
                    if not isinstance(r, list):
                        r = [r]
                    vals = [_coerce(v) for v in r]
                    ws.append(vals)
                    for i, v in enumerate(vals):
                        if i < len(widths):
                            widths[i] = max(widths[i], len(str(v)))
                        else:
                            widths.append(len(str(v)))

                for i, w in enumerate(widths):
                    ws.column_dimensions[get_column_letter(i + 1)].width = min(max(w + 2, 8), 60)

                if columns and rows:
                    last_col = get_column_letter(len(columns))
                    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

                chart_spec = sd.get("chart")
                if isinstance(chart_spec, dict) and columns and rows:
                    self._add_chart(ws, chart_spec, len(columns), len(rows))

            if not wb.sheetnames:
                wb.create_sheet("Sheet1")

            stored_name = new_stored_name("xlsx")
            path = os.path.join(generated_dir(), stored_name)
            wb.save(path)
        except Exception as exc:  # noqa: BLE001
            return ToolResult(content=f"Failed to build spreadsheet: {exc}", citations=None)

        download_name = safe_download_name(title, "xlsx", fallback="workbook")
        return download_result(
            stored_name, download_name, f"Created an Excel workbook with {count} sheet(s)."
        )
